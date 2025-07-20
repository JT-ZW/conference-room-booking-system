from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from utils.logging import log_user_activity
from utils.decorators import activity_logged
from core import supabase_admin, convert_datetime_strings, ActivityTypes
from datetime import datetime, UTC, timedelta, timezone
import io
import csv
import traceback
from collections import defaultdict
from decimal import Decimal
from decimal import Decimal

# CAT (Central Africa Time) timezone - UTC+2
CAT = timezone(timedelta(hours=2))

def get_current_time():
    """Get current time in CAT timezone (UTC+2)"""
    return datetime.now(CAT)

def get_utc_time():
    """Get current time in UTC for database operations"""
    return datetime.now(UTC)

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import xlsxwriter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

reports_bp = Blueprint('reports', __name__)

# Custom Jinja2 filters
@reports_bp.app_template_filter('as_date')
def as_date_filter(date_string):
    """Convert date string to date object"""
    try:
        if isinstance(date_string, str):
            return datetime.fromisoformat(date_string).date()
        return date_string
    except:
        return datetime.now().date()

# ===============================
# MAIN REPORTS DASHBOARD
# ===============================

@reports_bp.route('/reports')
@login_required
@activity_logged(ActivityTypes.GENERATE_REPORT, "Accessed reports dashboard")
def reports():
    """Reports dashboard with quick navigation to different report types"""
    try:
        print("🔍 DEBUG: Loading reports dashboard")
        
        # Get current date info for default values
        current_time = get_current_time()
        tomorrow = current_time + timedelta(days=1)
        next_week_start = tomorrow + timedelta(days=(7 - tomorrow.weekday()))
        next_month = current_time.replace(day=1) + timedelta(days=32)
        next_month = next_month.replace(day=1)
        
        dashboard_data = {
            'current_date': current_time.strftime('%Y-%m-%d'),
            'tomorrow_date': tomorrow.strftime('%Y-%m-%d'),
            'next_week_start': next_week_start.strftime('%Y-%m-%d'),
            'next_week_end': (next_week_start + timedelta(days=6)).strftime('%Y-%m-%d'),
            'next_month_year': next_month.year,
            'next_month_month': next_month.month,
            'next_month_name': next_month.strftime('%B %Y')
        }
        
        print(f"✅ DEBUG: Reports dashboard loaded")
        
        return render_template('reports/index.html', 
                             title='Reports Dashboard',
                             dashboard_data=dashboard_data,
                             moment=get_current_time)  # Pass timezone function
        
    except Exception as e:
        print(f"❌ ERROR: Failed to load reports dashboard: {e}")
        traceback.print_exc()
        
        flash('⚠️ Error loading reports dashboard. Please try again.', 'warning')
        return render_template('reports/index.html', 
                             title='Reports Dashboard',
                             dashboard_data={},
                             moment=get_current_time)

# ===============================
# DAILY SUMMARY REPORT
# ===============================

@reports_bp.route('/reports/daily-summary')
@login_required
@activity_logged(ActivityTypes.GENERATE_REPORT, "Generated daily summary report")
def daily_summary_report():
    """Generate daily summary report for tomorrow's events"""
    try:
        # Get date parameter (default to tomorrow in CAT timezone)
        date_str = request.args.get('date')
        if date_str:
            report_date = datetime.fromisoformat(date_str).date()
        else:
            # Default to tomorrow in CAT timezone
            tomorrow_cat = get_current_time() + timedelta(days=1)
            report_date = tomorrow_cat.date()
        
        print(f"🔍 DEBUG: Generating daily summary for {report_date}")
        
        # Get daily data for the specified date
        daily_data = get_daily_summary_data(report_date)
        
        return render_template('reports/daily_summary.html', 
                             title=f'Daily Summary - {report_date}',
                             daily_data=daily_data,
                             report_date=report_date,
                             moment=get_current_time,  # Pass timezone function
                             timedelta=timedelta)      # Pass timedelta for navigation
        
    except Exception as e:
        print(f"❌ ERROR: Failed to generate daily summary: {e}")
        traceback.print_exc()
        flash('❌ Error generating daily summary. Please try again.', 'danger')
        
        # Return with empty data structure to prevent template errors
        empty_data = {
            'date': report_date,
            'events_by_room': {},
            'summary': {
                'total_events': 0,
                'confirmed_events': 0,
                'tentative_events': 0,
                'cancelled_events': 0,
                'total_revenue': 0,
                'total_attendees': 0,
                'rooms_in_use': 0,
                'average_event_value': 0
            },
            'bookings': []
        }
        
        return render_template('reports/daily_summary.html', 
                             title=f'Daily Summary - {report_date}',
                             daily_data=empty_data,
                             report_date=report_date,
                             moment=get_current_time,
                             timedelta=timedelta)

@reports_bp.route('/reports/daily-summary/export')
@login_required
def export_daily_summary():
    """Export daily summary report"""
    try:
        # Get parameters
        date_str = request.args.get('date')
        format_type = request.args.get('format', 'excel')
        
        if date_str:
            report_date = datetime.fromisoformat(date_str).date()
        else:
            report_date = (get_current_time() + timedelta(days=1)).date()
        
        # Get daily data
        daily_data = get_daily_summary_data(report_date)
        
        if format_type == 'pdf' and REPORTLAB_AVAILABLE:
            return export_daily_summary_pdf(daily_data, report_date)
        else:
            return export_daily_summary_excel(daily_data, report_date)
            
    except Exception as e:
        print(f"❌ ERROR: Failed to export daily summary: {e}")
        flash('❌ Error exporting daily summary. Please try again.', 'danger')
        return redirect(url_for('reports.daily_summary_report'))

# ===============================
# WEEKLY SUMMARY REPORT
# ===============================

@reports_bp.route('/reports/weekly-summary')
@login_required
@activity_logged(ActivityTypes.GENERATE_REPORT, "Generated weekly summary report")
def weekly_summary_report():
    """Generate weekly summary report for next week's events in table format"""
    try:
        # Get week parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            # Default to next week (starting from tomorrow in CAT)
            current_time = get_current_time()
            tomorrow = current_time + timedelta(days=1)
            
            # Find the start of next week (Monday)
            days_until_monday = (7 - tomorrow.weekday()) % 7
            if days_until_monday == 0:  # If tomorrow is Monday
                start_of_week = tomorrow
            else:
                start_of_week = tomorrow + timedelta(days=days_until_monday)
            
            end_of_week = start_of_week + timedelta(days=6)
            start_date = start_of_week.strftime('%Y-%m-%d')
            end_date = end_of_week.strftime('%Y-%m-%d')
        
        print(f"🔍 DEBUG: Generating weekly summary for {start_date} to {end_date}")
        
        # Get weekly data in table format
        weekly_data = get_weekly_summary_data(start_date, end_date)
        
        return render_template('reports/weekly_summary.html', 
                             title=f'Weekly Summary - {start_date} to {end_date}',
                             weekly_data=weekly_data,
                             start_date=start_date,
                             end_date=end_date,
                             moment=get_current_time,  # Pass timezone function
                             timedelta=timedelta)      # Pass timedelta for navigation
        
    except Exception as e:
        print(f"❌ ERROR: Failed to generate weekly summary: {e}")
        traceback.print_exc()
        flash('❌ Error generating weekly summary. Please try again.', 'danger')
        
        # Return with empty data structure to prevent template errors
        empty_data = {
            'start_date': start_date,
            'end_date': end_date,
            'week_days': [],
            'room_schedule': {},
            'summary': {
                'total_events': 0,
                'total_revenue': 0,
                'total_attendees': 0,
                'rooms_with_events': 0,
                'average_daily_events': 0
            },
            'bookings': []
        }
        
        return render_template('reports/weekly_summary.html', 
                             title=f'Weekly Summary - {start_date} to {end_date}',
                             weekly_data=empty_data,
                             start_date=start_date,
                             end_date=end_date,
                             moment=get_current_time,
                             timedelta=timedelta)

@reports_bp.route('/reports/weekly-summary/export')
@login_required
def export_weekly_summary():
    """Export weekly summary report"""
    try:
        # Get parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        format_type = request.args.get('format', 'excel')
        
        if not start_date or not end_date:
            current_time = get_current_time()
            tomorrow = current_time + timedelta(days=1)
            days_until_monday = (7 - tomorrow.weekday()) % 7
            if days_until_monday == 0:
                start_of_week = tomorrow
            else:
                start_of_week = tomorrow + timedelta(days=days_until_monday)
            end_of_week = start_of_week + timedelta(days=6)
            start_date = start_of_week.strftime('%Y-%m-%d')
            end_date = end_of_week.strftime('%Y-%m-%d')
        
        # Get weekly data
        weekly_data = get_weekly_summary_data(start_date, end_date)
        
        if format_type == 'pdf' and REPORTLAB_AVAILABLE:
            return export_weekly_summary_pdf(weekly_data, start_date, end_date)
        else:
            return export_weekly_summary_excel(weekly_data, start_date, end_date)
            
    except Exception as e:
        print(f"❌ ERROR: Failed to export weekly summary: {e}")
        flash('❌ Error exporting weekly summary. Please try again.', 'danger')
        return redirect(url_for('reports.weekly_summary_report'))

# ===============================
# MONTHLY SUMMARY REPORT
# ===============================

@reports_bp.route('/reports/monthly-summary')
@login_required
@activity_logged(ActivityTypes.GENERATE_REPORT, "Generated monthly summary report")
def monthly_summary_report():
    """Generate monthly summary report for next month's events"""
    try:
        # Get month parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not year or not month:
            # Default to next month in CAT timezone
            current_time = get_current_time()
            next_month_date = current_time.replace(day=1) + timedelta(days=32)
            year = next_month_date.year
            month = next_month_date.month
        
        print(f"🔍 DEBUG: Generating monthly summary for {year}-{month:02d}")
        
        # Calculate month date range
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        # Get monthly data
        monthly_data = get_monthly_summary_data(start_date, end_date)
        
        return render_template('reports/monthly_summary.html', 
                             title=f'Monthly Summary - {start_date.strftime("%B %Y")}',
                             monthly_data=monthly_data,
                             year=year,
                             month=month,
                             start_date=start_date,
                             end_date=end_date,
                             moment=get_current_time,  # Pass timezone function
                             timedelta=timedelta)      # Pass timedelta for navigation
        
    except Exception as e:
        print(f"❌ ERROR: Failed to generate monthly summary: {e}")
        traceback.print_exc()
        flash('❌ Error generating monthly summary. Please try again.', 'danger')
        return redirect(url_for('reports.reports'))

@reports_bp.route('/reports/monthly-summary/export')
@login_required
def export_monthly_summary():
    """Export monthly summary report"""
    try:
        # Get parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        format_type = request.args.get('format', 'excel')
        
        if not year or not month:
            current_time = get_current_time()
            next_month_date = current_time.replace(day=1) + timedelta(days=32)
            year = next_month_date.year
            month = next_month_date.month
        
        # Calculate month date range
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        # Get monthly data
        monthly_data = get_monthly_summary_data(start_date, end_date)
        
        if format_type == 'pdf' and REPORTLAB_AVAILABLE:
            return export_monthly_summary_pdf(monthly_data, start_date, end_date)
        else:
            return export_monthly_summary_excel(monthly_data, start_date, end_date)
            
    except Exception as e:
        print(f"❌ ERROR: Failed to export monthly summary: {e}")
        flash('❌ Error exporting monthly summary. Please try again.', 'danger')
        return redirect(url_for('reports.monthly_summary_report'))

# ===============================
# DATA RETRIEVAL FUNCTIONS
# ===============================

def get_daily_summary_data(report_date):
    """Get comprehensive daily summary data for the specified date"""
    try:
        # Convert date to datetime range in UTC for database query
        start_dt = datetime.combine(report_date, datetime.min.time()).replace(tzinfo=UTC)
        end_dt = datetime.combine(report_date, datetime.max.time()).replace(tzinfo=UTC)
        
        print(f"🔍 DEBUG: Fetching bookings for {report_date} ({start_dt} to {end_dt})")
        
        # Get bookings that START on the specified date
        response = supabase_admin.table('bookings').select("""
            *,
            room:rooms(id, name, capacity),
            client:clients(id, contact_person, company_name, email, phone)
        """).gte('start_time', start_dt.isoformat()).lte(
            'start_time', end_dt.isoformat()
        ).order('start_time').execute()
        
        bookings = response.data if response.data else []
        print(f"🔍 DEBUG: Found {len(bookings)} bookings for {report_date}")
        
        # Convert datetime strings
        for booking in bookings:
            booking = convert_datetime_strings(booking)
        
        # Group events by room for organized display
        events_by_room = defaultdict(list)
        total_events = 0
        total_confirmed = 0
        total_tentative = 0
        total_cancelled = 0
        total_revenue = 0
        total_attendees = 0
        
        for booking in bookings:
            status = booking.get('status', 'tentative')
            
            # Count by status
            if status == 'confirmed':
                total_confirmed += 1
                total_revenue += float(booking.get('total_price', 0))
            elif status == 'tentative':
                total_tentative += 1
            elif status == 'cancelled':
                total_cancelled += 1
                continue  # Don't include cancelled in room grouping
            
            total_events += 1
            total_attendees += int(booking.get('attendees', 0))
            
            # Get room information
            room_info = booking.get('room', {})
            room_name = room_info.get('name', 'Unknown Room')
            
            # Format event details for display
            event_details = format_event_details(booking)
            events_by_room[room_name].append(event_details)
        
        # Sort rooms alphabetically
        events_by_room = dict(sorted(events_by_room.items()))
        
        # Calculate summary statistics
        summary = {
            'total_events': total_events,
            'confirmed_events': total_confirmed,
            'tentative_events': total_tentative,
            'cancelled_events': total_cancelled,
            'total_revenue': round(total_revenue, 2),
            'total_attendees': total_attendees,
            'rooms_in_use': len(events_by_room),
            'average_event_value': round(total_revenue / max(total_confirmed, 1), 2)
        }
        
        return {
            'date': report_date,
            'events_by_room': events_by_room,
            'summary': summary,
            'bookings': bookings  # Keep original bookings for export
        }
        
    except Exception as e:
        print(f"❌ ERROR: Failed to get daily summary: {e}")
        traceback.print_exc()
        return {
            'date': report_date,
            'events_by_room': {},
            'summary': {
                'total_events': 0,
                'confirmed_events': 0,
                'tentative_events': 0,
                'cancelled_events': 0,
                'total_revenue': 0,
                'total_attendees': 0,
                'rooms_in_use': 0,
                'average_event_value': 0
            },
            'bookings': []
        }

def get_weekly_summary_data(start_date, end_date):
    """Get weekly summary data in table format (rooms x days)"""
    try:
        # Convert to datetime range for database query
        start_dt = datetime.fromisoformat(start_date).replace(hour=0, minute=0, second=0, tzinfo=UTC)
        end_dt = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59, tzinfo=UTC)
        
        print(f"🔍 DEBUG: Fetching weekly bookings from {start_dt} to {end_dt}")
        
        # Get all bookings for the week
        response = supabase_admin.table('bookings').select("""
            *,
            room:rooms(id, name, capacity),
            client:clients(id, contact_person, company_name, email, phone)
        """).gte('start_time', start_dt.isoformat()).lte(
            'start_time', end_dt.isoformat()
        ).neq('status', 'cancelled').order('start_time').execute()
        
        bookings = response.data if response.data else []
        print(f"🔍 DEBUG: Found {len(bookings)} bookings for the week")
        
        # Convert datetime strings
        for booking in bookings:
            booking = convert_datetime_strings(booking)
        
        # Get all rooms for the table structure
        rooms_response = supabase_admin.table('rooms').select('*').order('name').execute()
        rooms = rooms_response.data if rooms_response.data else []
        
        # Create week structure
        week_days = []
        current_date = datetime.fromisoformat(start_date).date()
        end_date_obj = datetime.fromisoformat(end_date).date()
        
        while current_date <= end_date_obj:
            week_days.append({
                'date': current_date.strftime('%Y-%m-%d'),  # Convert to string for JSON
                'date_obj': current_date,  # Keep original for internal use
                'day_name': current_date.strftime('%A'),
                'date_display': current_date.strftime('%d/%m/%y')
            })
            current_date += timedelta(days=1)
        
        # Create room schedule grid
        room_schedule = {}
        
        for room in rooms:
            room_id = room['id']
            room_name = room['name']
            
            room_schedule[room_name] = {
                'room_info': room,
                'days': {}
            }
            
            # Initialize each day for this room
            for day_info in week_days:
                # Use string representation of date as key for JSON serialization
                date_str = day_info['date']  # Already a string now
                room_schedule[room_name]['days'][date_str] = []
        
        # Populate schedule with bookings
        total_events = 0
        total_revenue = 0
        total_attendees = 0
        
        for booking in bookings:
            if booking.get('start_time') and booking.get('room'):
                start_time = booking['start_time']
                if isinstance(start_time, str):
                    start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                
                booking_date = start_time.date()
                booking_date_str = booking_date.strftime('%Y-%m-%d')  # Convert to string
                room_name = booking['room'].get('name', 'Unknown Room')
                
                if room_name in room_schedule and booking_date_str in room_schedule[room_name]['days']:
                    event_details = format_event_details_for_table(booking)
                    room_schedule[room_name]['days'][booking_date_str].append(event_details)
                    
                    # Update totals
                    total_events += 1
                    if booking.get('status') == 'confirmed':
                        total_revenue += float(booking.get('total_price', 0))
                    total_attendees += int(booking.get('attendees', 0))
        
        # Calculate summary
        summary = {
            'total_events': total_events,
            'total_revenue': round(total_revenue, 2),
            'total_attendees': total_attendees,
            'rooms_with_events': len([r for r in room_schedule.values() 
                                    if any(day_events for day_events in r['days'].values())]),
            'average_daily_events': round(total_events / 7, 1)
        }
        
        return {
            'start_date': start_date,
            'end_date': end_date,
            'week_days': week_days,
            'room_schedule': room_schedule,
            'summary': summary,
            'bookings': bookings  # Keep original bookings for export
        }
        
    except Exception as e:
        print(f"❌ ERROR: Failed to get weekly summary: {e}")
        traceback.print_exc()
        return {
            'start_date': start_date,
            'end_date': end_date,
            'week_days': [],
            'room_schedule': {},
            'summary': {
                'total_events': 0,
                'total_revenue': 0,
                'total_attendees': 0,
                'rooms_with_events': 0,
                'average_daily_events': 0
            },
            'bookings': []
        }

def get_monthly_summary_data(start_date, end_date):
    """Get monthly summary data with weekly breakdown"""
    try:
        # Convert to datetime range for database query
        start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=UTC)
        end_dt = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=UTC)
        
        print(f"🔍 DEBUG: Fetching monthly bookings from {start_dt} to {end_dt}")
        
        # Get all bookings for the month
        response = supabase_admin.table('bookings').select("""
            *,
            room:rooms(id, name, capacity),
            client:clients(id, contact_person, company_name, email, phone)
        """).gte('start_time', start_dt.isoformat()).lte(
            'start_time', end_dt.isoformat()
        ).order('start_time').execute()
        
        bookings = response.data if response.data else []
        print(f"🔍 DEBUG: Found {len(bookings)} bookings for the month")
        
        # Convert datetime strings
        for booking in bookings:
            booking = convert_datetime_strings(booking)
        
        # Group bookings by week
        weekly_breakdown = defaultdict(list)
        weekly_summaries = {}
        
        # Calculate weeks in the month
        current_date = start_date
        week_number = 1
        
        while current_date <= end_date:
            week_start = current_date
            week_end = min(current_date + timedelta(days=6), end_date)
            
            week_key = f"Week {week_number}"
            weekly_summaries[week_key] = {
                'start_date': week_start,
                'end_date': week_end,
                'events': 0,
                'confirmed': 0,
                'tentative': 0,
                'cancelled': 0,
                'revenue': 0,
                'attendees': 0
            }
            
            current_date = week_end + timedelta(days=1)
            week_number += 1
        
        # Process bookings
        total_events = 0
        total_confirmed = 0
        total_tentative = 0
        total_cancelled = 0
        total_revenue = 0
        total_attendees = 0
        
        # Group by room for room utilization
        room_utilization = defaultdict(lambda: {
            'events': 0,
            'confirmed': 0,
            'revenue': 0,
            'attendees': 0
        })
        
        # Group by client for client analysis
        client_analysis = defaultdict(lambda: {
            'events': 0,
            'confirmed': 0,
            'revenue': 0,
            'attendees': 0
        })
        
        for booking in bookings:
            start_time = booking.get('start_time')
            if isinstance(start_time, str):
                start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            
            booking_date = start_time.date()
            status = booking.get('status', 'tentative')
            revenue = float(booking.get('total_price', 0))
            attendees = int(booking.get('attendees', 0))
            
            # Update totals
            total_events += 1
            total_attendees += attendees
            
            if status == 'confirmed':
                total_confirmed += 1
                total_revenue += revenue
            elif status == 'tentative':
                total_tentative += 1
            elif status == 'cancelled':
                total_cancelled += 1
            
            # Find which week this booking belongs to
            for week_key, week_info in weekly_summaries.items():
                if week_info['start_date'] <= booking_date <= week_info['end_date']:
                    week_info['events'] += 1
                    week_info['attendees'] += attendees
                    
                    if status == 'confirmed':
                        week_info['confirmed'] += 1
                        week_info['revenue'] += revenue
                    elif status == 'tentative':
                        week_info['tentative'] += 1
                    elif status == 'cancelled':
                        week_info['cancelled'] += 1
                    break
            
            # Room utilization
            room_name = booking.get('room', {}).get('name', 'Unknown Room')
            room_utilization[room_name]['events'] += 1
            room_utilization[room_name]['attendees'] += attendees
            if status == 'confirmed':
                room_utilization[room_name]['confirmed'] += 1
                room_utilization[room_name]['revenue'] += revenue
            
            # Client analysis
            client = booking.get('client', {})
            client_name = client.get('company_name') or client.get('contact_person', 'Unknown Client')
            client_analysis[client_name]['events'] += 1
            client_analysis[client_name]['attendees'] += attendees
            if status == 'confirmed':
                client_analysis[client_name]['confirmed'] += 1
                client_analysis[client_name]['revenue'] += revenue
        
        # Convert to sorted lists for display
        top_rooms = sorted(room_utilization.items(), key=lambda x: x[1]['revenue'], reverse=True)[:10]
        top_clients = sorted(client_analysis.items(), key=lambda x: x[1]['revenue'], reverse=True)[:10]
        
        # Overall summary
        summary = {
            'total_events': total_events,
            'confirmed_events': total_confirmed,
            'tentative_events': total_tentative,
            'cancelled_events': total_cancelled,
            'total_revenue': round(total_revenue, 2),
            'total_attendees': total_attendees,
            'conversion_rate': round((total_confirmed / max(total_events, 1)) * 100, 1),
            'average_event_value': round(total_revenue / max(total_confirmed, 1), 2),
            'days_in_month': (end_date - start_date).days + 1
        }
        
        return {
            'start_date': start_date,
            'end_date': end_date,
            'month_name': start_date.strftime('%B %Y'),
            'weekly_summaries': weekly_summaries,
            'summary': summary,
            'top_rooms': top_rooms,
            'top_clients': top_clients,
            'bookings': bookings  # Keep original bookings for export
        }
        
    except Exception as e:
        print(f"❌ ERROR: Failed to get monthly summary: {e}")
        traceback.print_exc()
        return {
            'start_date': start_date,
            'end_date': end_date,
            'month_name': start_date.strftime('%B %Y'),
            'weekly_summaries': {},
            'summary': {
                'total_events': 0,
                'confirmed_events': 0,
                'tentative_events': 0,
                'cancelled_events': 0,
                'total_revenue': 0,
                'total_attendees': 0,
                'conversion_rate': 0,
                'average_event_value': 0,
                'days_in_month': 0
            },
            'top_rooms': [],
            'top_clients': [],
            'bookings': []
        }

# ===============================
# FORMATTING HELPER FUNCTIONS
# ===============================

def format_event_details(booking):
    """Format event details for daily summary display"""
    try:
        # Get basic information
        title = booking.get('title', 'Event')
        attendees = booking.get('attendees', 0)
        status = booking.get('status', 'tentative')
        
        # Get client information
        client = booking.get('client', {})
        client_name = client.get('company_name') or client.get('contact_person', 'Unknown Client')
        
        # Get time information
        start_time = booking.get('start_time')
        end_time = booking.get('end_time')
        
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
        if isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
        
        time_display = f"{start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}"
        duration = end_time - start_time
        duration_hours = duration.total_seconds() / 3600
        
        # Get pricing information
        total_price = float(booking.get('total_price', 0))
        price_per_person = total_price / max(attendees, 1) if attendees > 0 else 0
        
        return {
            'id': booking.get('id'),
            'title': title,
            'client_name': client_name,
            'attendees': attendees,
            'time_display': time_display,
            'duration_hours': round(duration_hours, 1),
            'total_price': round(total_price, 2),
            'price_per_person': round(price_per_person, 2),
            'status': status,
            'status_display': status.replace('_', ' ').title(),
            'notes': booking.get('notes', ''),
            'start_time': start_time,
            'end_time': end_time,
            'raw_booking': booking  # Include full booking data for template access
        }
        
    except Exception as e:
        print(f"⚠️ ERROR: Failed to format event details: {e}")
        return {
            'id': booking.get('id', 'unknown'),
            'title': 'Event Details Loading...',
            'client_name': 'Client Loading...',
            'attendees': 0,
            'time_display': 'Time TBD',
            'duration_hours': 0,
            'total_price': 0,
            'price_per_person': 0,
            'status': 'unknown',
            'status_display': 'Loading...',
            'notes': ''
        }

def format_event_details_for_table(booking):
    """Format event details for weekly table display (condensed format)"""
    try:
        # Get basic information
        client = booking.get('client', {})
        client_name = client.get('company_name') or client.get('contact_person', 'Unknown Client')
        
        # Shorten long client names for table display
        if len(client_name) > 20:
            client_name = client_name[:17] + "..."
        
        attendees = booking.get('attendees', 0)
        
        # Get time information
        start_time = booking.get('start_time')
        end_time = booking.get('end_time')
        
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
        if isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
        
        duration = end_time - start_time
        duration_hours = duration.total_seconds() / 3600
        
        # Get pricing information
        total_price = float(booking.get('total_price', 0))
        price_per_person = total_price / max(attendees, 1) if attendees > 0 else 0
        
        # Format for table cell (similar to your document format)
        formatted_text = f"**{client_name}**\n"
        formatted_text += f"**{attendees}PAX {start_time.strftime('%H%M')}HOURS**\n"
        formatted_text += f"**${price_per_person:.0f}PP**"
        
        return {
            'id': booking.get('id'),
            'client_name': client_name,
            'attendees': attendees,
            'start_time': start_time.strftime('%H%M'),
            'duration_hours': round(duration_hours, 1),
            'price_per_person': round(price_per_person, 2),
            'total_price': round(total_price, 2),
            'status': booking.get('status', 'tentative'),
            'formatted_text': formatted_text,
            'raw_booking': booking  # Include full booking data for template access
        }
        
    except Exception as e:
        print(f"⚠️ ERROR: Failed to format table event details: {e}")
        return {
            'id': booking.get('id', 'unknown'),
            'client_name': 'Loading...',
            'attendees': 0,
            'start_time': 'TBD',
            'duration_hours': 0,
            'price_per_person': 0,
            'total_price': 0,
            'status': 'unknown',
            'formatted_text': '**Loading...**',
            'raw_booking': booking
        }

# ===============================
# EXPORT FUNCTIONS - EXCEL
# ===============================

def export_daily_summary_excel(daily_data, report_date):
    """Export daily summary as Excel file"""
    try:
        if not EXCEL_AVAILABLE:
            return export_daily_summary_csv(daily_data, report_date)
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        # Create formats
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'bg_color': '#2E86AB',
            'font_color': 'white',
            'border': 1
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#A23B72',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })
        
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'top',
            'text_wrap': True
        })
        
        # Create worksheet
        worksheet = workbook.add_worksheet('Daily Summary')
        
        # Set column widths
        worksheet.set_column('A:A', 20)  # Room
        worksheet.set_column('B:B', 25)  # Event
        worksheet.set_column('C:C', 20)  # Client
        worksheet.set_column('D:D', 10)  # PAX
        worksheet.set_column('E:E', 15)  # Time
        worksheet.set_column('F:F', 10)  # Duration
        worksheet.set_column('G:G', 12)  # Revenue
        worksheet.set_column('H:H', 10)  # Status
        
        # Title
        worksheet.merge_range('A1:H1', f'Daily Summary Report - {report_date.strftime("%A, %B %d, %Y")}', title_format)
        
        # Summary statistics
        row = 3
        summary = daily_data['summary']
        worksheet.write(row, 0, 'SUMMARY STATISTICS', header_format)
        worksheet.merge_range(row, 1, row, 7, '', header_format)
        row += 1
        
        stats = [
            ['Total Events', summary['total_events']],
            ['Confirmed Events', summary['confirmed_events']],
            ['Tentative Events', summary['tentative_events']],
            ['Total Revenue', f"${summary['total_revenue']:,.2f}"],
            ['Total Attendees', f"{summary['total_attendees']:,}"],
            ['Rooms in Use', summary['rooms_in_use']]
        ]
        
        for stat_name, stat_value in stats:
            worksheet.write(row, 0, stat_name, cell_format)
            worksheet.write(row, 1, stat_value, cell_format)
            row += 1
        
        # Events by room
        row += 2
        worksheet.write(row, 0, 'EVENTS BY ROOM', header_format)
        worksheet.merge_range(row, 1, row, 7, '', header_format)
        row += 1
        
        # Headers
        headers = ['Room', 'Event', 'Client', 'PAX', 'Time', 'Duration (hrs)', 'Revenue', 'Status']
        for col, header in enumerate(headers):
            worksheet.write(row, col, header, header_format)
        row += 1
        
        # Events data
        for room_name, events in daily_data['events_by_room'].items():
            for i, event in enumerate(events):
                # Room name only on first event for this room
                if i == 0:
                    worksheet.write(row, 0, room_name, cell_format)
                else:
                    worksheet.write(row, 0, '', cell_format)
                
                worksheet.write(row, 1, event['title'], cell_format)
                worksheet.write(row, 2, event['client_name'], cell_format)
                worksheet.write(row, 3, event['attendees'], cell_format)
                worksheet.write(row, 4, event['time_display'], cell_format)
                worksheet.write(row, 5, event['duration_hours'], cell_format)
                worksheet.write(row, 6, f"${event['total_price']:,.2f}", cell_format)
                worksheet.write(row, 7, event['status_display'], cell_format)
                row += 1
            
            # Add empty row between rooms
            if events:
                row += 1
        
        workbook.close()
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'daily_summary_{report_date.strftime("%Y-%m-%d")}.xlsx'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export daily summary Excel: {e}")
        return export_daily_summary_csv(daily_data, report_date)

def export_weekly_summary_excel(weekly_data, start_date, end_date):
    """Export weekly summary as Excel file in table format"""
    try:
        if not EXCEL_AVAILABLE:
            return export_weekly_summary_csv(weekly_data, start_date, end_date)
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        # Create formats
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'bg_color': '#2E86AB',
            'font_color': 'white',
            'border': 1
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#A23B72',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'text_wrap': True
        })
        
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'top',
            'text_wrap': True,
            'font_size': 9
        })
        
        # Create worksheet
        worksheet = workbook.add_worksheet('Weekly Summary')
        
        # Set column widths
        worksheet.set_column('A:A', 15)  # Room column
        for i in range(len(weekly_data['week_days'])):
            worksheet.set_column(i + 1, i + 1, 18)  # Day columns
        
        # Set row height for better text wrapping
        worksheet.set_default_row(60)
        
        # Title
        title_text = f'Weekly Summary Report - {start_date} to {end_date}'
        worksheet.merge_range(0, 0, 0, len(weekly_data['week_days']), title_text, title_format)
        
        # Headers
        row = 2
        worksheet.write(row, 0, 'ROOM', header_format)
        
        for col, day in enumerate(weekly_data['week_days']):
            day_header = f"{day['day_name']}\n{day['date_display']}"
            worksheet.write(row, col + 1, day_header, header_format)
        
        row += 1
        
        # Room schedule data
        for room_name, room_data in weekly_data['room_schedule'].items():
            worksheet.write(row, 0, room_name, cell_format)
            
            for col, day in enumerate(weekly_data['week_days']):
                day_date = day['date']
                events = room_data['days'].get(day_date, [])
                
                if events:
                    cell_content = '\n\n'.join([
                        f"{event['client_name']}\n{event['attendees']}PAX {event['start_time']}\n${event['price_per_person']:.0f}PP"
                        for event in events
                    ])
                else:
                    cell_content = ''
                
                worksheet.write(row, col + 1, cell_content, cell_format)
            
            row += 1
        
        # Summary section
        row += 2
        summary = weekly_data['summary']
        worksheet.write(row, 0, 'WEEKLY SUMMARY', header_format)
        row += 1
        
        summary_stats = [
            ['Total Events', summary['total_events']],
            ['Total Revenue', f"${summary['total_revenue']:,.2f}"],
            ['Total Attendees', f"{summary['total_attendees']:,}"],
            ['Rooms with Events', summary['rooms_with_events']],
            ['Average Daily Events', summary['average_daily_events']]
        ]
        
        for stat_name, stat_value in summary_stats:
            worksheet.write(row, 0, stat_name, cell_format)
            worksheet.write(row, 1, stat_value, cell_format)
            row += 1
        
        workbook.close()
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'weekly_summary_{start_date}_to_{end_date}.xlsx'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export weekly summary Excel: {e}")
        return export_weekly_summary_csv(weekly_data, start_date, end_date)

def export_monthly_summary_excel(monthly_data, start_date, end_date):
    """Export monthly summary as Excel file"""
    try:
        if not EXCEL_AVAILABLE:
            return export_monthly_summary_csv(monthly_data, start_date, end_date)
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        # Create formats
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'bg_color': '#2E86AB',
            'font_color': 'white',
            'border': 1
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#A23B72',
            'font_color': 'white',
            'border': 1,
            'align': 'center'
        })
        
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left'
        })
        
        # Create worksheet
        worksheet = workbook.add_worksheet('Monthly Summary')
        
        # Set column widths
        worksheet.set_column('A:H', 15)
        
        # Title
        worksheet.merge_range('A1:H1', f'Monthly Summary Report - {monthly_data["month_name"]}', title_format)
        
        # Overall summary
        row = 3
        summary = monthly_data['summary']
        worksheet.write(row, 0, 'OVERALL SUMMARY', header_format)
        worksheet.merge_range(row, 1, row, 7, '', header_format)
        row += 1
        
        overall_stats = [
            ['Total Events', summary['total_events']],
            ['Confirmed Events', summary['confirmed_events']],
            ['Tentative Events', summary['tentative_events']],
            ['Total Revenue', f"${summary['total_revenue']:,.2f}"],
            ['Total Attendees', f"{summary['total_attendees']:,}"],
            ['Conversion Rate', f"{summary['conversion_rate']}%"],
            ['Average Event Value', f"${summary['average_event_value']:,.2f}"]
        ]
        
        for stat_name, stat_value in overall_stats:
            worksheet.write(row, 0, stat_name, cell_format)
            worksheet.write(row, 1, stat_value, cell_format)
            row += 1
        
        # Weekly breakdown
        row += 2
        worksheet.write(row, 0, 'WEEKLY BREAKDOWN', header_format)
        worksheet.merge_range(row, 1, row, 7, '', header_format)
        row += 1
        
        # Weekly headers
        weekly_headers = ['Week', 'Events', 'Confirmed', 'Tentative', 'Revenue', 'Attendees']
        for col, header in enumerate(weekly_headers):
            worksheet.write(row, col, header, header_format)
        row += 1
        
        # Weekly data
        for week_name, week_data in monthly_data['weekly_summaries'].items():
            worksheet.write(row, 0, week_name, cell_format)
            worksheet.write(row, 1, week_data['events'], cell_format)
            worksheet.write(row, 2, week_data['confirmed'], cell_format)
            worksheet.write(row, 3, week_data['tentative'], cell_format)
            worksheet.write(row, 4, f"${week_data['revenue']:,.2f}", cell_format)
            worksheet.write(row, 5, week_data['attendees'], cell_format)
            row += 1
        
        # Top rooms
        row += 2
        worksheet.write(row, 0, 'TOP PERFORMING ROOMS', header_format)
        worksheet.merge_range(row, 1, row, 7, '', header_format)
        row += 1
        
        room_headers = ['Room', 'Events', 'Confirmed', 'Revenue', 'Attendees']
        for col, header in enumerate(room_headers):
            worksheet.write(row, col, header, header_format)
        row += 1
        
        for room_name, room_data in monthly_data['top_rooms'][:10]:
            worksheet.write(row, 0, room_name, cell_format)
            worksheet.write(row, 1, room_data['events'], cell_format)
            worksheet.write(row, 2, room_data['confirmed'], cell_format)
            worksheet.write(row, 3, f"${room_data['revenue']:,.2f}", cell_format)
            worksheet.write(row, 4, room_data['attendees'], cell_format)
            row += 1
        
        # Top clients
        row += 2
        worksheet.write(row, 0, 'TOP CLIENTS', header_format)
        worksheet.merge_range(row, 1, row, 7, '', header_format)
        row += 1
        
        client_headers = ['Client', 'Events', 'Confirmed', 'Revenue', 'Attendees']
        for col, header in enumerate(client_headers):
            worksheet.write(row, col, header, header_format)
        row += 1
        
        for client_name, client_data in monthly_data['top_clients'][:10]:
            worksheet.write(row, 0, client_name, cell_format)
            worksheet.write(row, 1, client_data['events'], cell_format)
            worksheet.write(row, 2, client_data['confirmed'], cell_format)
            worksheet.write(row, 3, f"${client_data['revenue']:,.2f}", cell_format)
            worksheet.write(row, 4, client_data['attendees'], cell_format)
            row += 1
        
        workbook.close()
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'monthly_summary_{start_date.strftime("%Y-%m")}.xlsx'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export monthly summary Excel: {e}")
        return export_monthly_summary_csv(monthly_data, start_date, end_date)

# ===============================
# EXPORT FUNCTIONS - CSV FALLBACK
# ===============================

def export_daily_summary_csv(daily_data, report_date):
    """Export daily summary as CSV file (fallback)"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([f'Daily Summary Report - {report_date.strftime("%A, %B %d, %Y")}'])
        writer.writerow([f'Generated: {get_current_time().strftime("%Y-%m-%d %H:%M CAT")}'])
        writer.writerow([])
        
        # Summary
        summary = daily_data['summary']
        writer.writerow(['SUMMARY STATISTICS'])
        writer.writerow(['Total Events', summary['total_events']])
        writer.writerow(['Confirmed Events', summary['confirmed_events']])
        writer.writerow(['Tentative Events', summary['tentative_events']])
        writer.writerow(['Total Revenue', f"${summary['total_revenue']:,.2f}"])
        writer.writerow(['Total Attendees', f"{summary['total_attendees']:,}"])
        writer.writerow(['Rooms in Use', summary['rooms_in_use']])
        writer.writerow([])
        
        # Events
        writer.writerow(['EVENTS BY ROOM'])
        writer.writerow(['Room', 'Event', 'Client', 'PAX', 'Time', 'Duration (hrs)', 'Revenue', 'Status'])
        
        for room_name, events in daily_data['events_by_room'].items():
            for event in events:
                writer.writerow([
                    room_name,
                    event['title'],
                    event['client_name'],
                    event['attendees'],
                    event['time_display'],
                    event['duration_hours'],
                    f"${event['total_price']:,.2f}",
                    event['status_display']
                ])
        
        output.seek(0)
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'daily_summary_{report_date.strftime("%Y-%m-%d")}.csv'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export daily summary CSV: {e}")
        raise

def export_weekly_summary_csv(weekly_data, start_date, end_date):
    """Export weekly summary as CSV file (fallback)"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([f'Weekly Summary Report - {start_date} to {end_date}'])
        writer.writerow([f'Generated: {get_current_time().strftime("%Y-%m-%d %H:%M CAT")}'])
        writer.writerow([])
        
        # Create table structure
        headers = ['Room'] + [f"{day['day_name']} {day['date_display']}" for day in weekly_data['week_days']]
        writer.writerow(headers)
        
        for room_name, room_data in weekly_data['room_schedule'].items():
            row = [room_name]
            
            for day in weekly_data['week_days']:
                day_date = day['date']
                events = room_data['days'].get(day_date, [])
                
                if events:
                    cell_content = ' | '.join([
                        f"{event['client_name']} {event['attendees']}PAX {event['start_time']} ${event['price_per_person']:.0f}PP"
                        for event in events
                    ])
                else:
                    cell_content = ''
                
                row.append(cell_content)
            
            writer.writerow(row)
        
        # Summary
        writer.writerow([])
        writer.writerow(['WEEKLY SUMMARY'])
        summary = weekly_data['summary']
        writer.writerow(['Total Events', summary['total_events']])
        writer.writerow(['Total Revenue', f"${summary['total_revenue']:,.2f}"])
        writer.writerow(['Total Attendees', f"{summary['total_attendees']:,}"])
        writer.writerow(['Rooms with Events', summary['rooms_with_events']])
        
        output.seek(0)
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'weekly_summary_{start_date}_to_{end_date}.csv'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export weekly summary CSV: {e}")
        raise

def export_monthly_summary_csv(monthly_data, start_date, end_date):
    """Export monthly summary as CSV file (fallback)"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([f'Monthly Summary Report - {monthly_data["month_name"]}'])
        writer.writerow([f'Generated: {get_current_time().strftime("%Y-%m-%d %H:%M CAT")}'])
        writer.writerow([])
        
        # Overall summary
        summary = monthly_data['summary']
        writer.writerow(['OVERALL SUMMARY'])
        writer.writerow(['Total Events', summary['total_events']])
        writer.writerow(['Confirmed Events', summary['confirmed_events']])
        writer.writerow(['Tentative Events', summary['tentative_events']])
        writer.writerow(['Total Revenue', f"${summary['total_revenue']:,.2f}"])
        writer.writerow(['Total Attendees', f"{summary['total_attendees']:,}"])
        writer.writerow(['Conversion Rate', f"{summary['conversion_rate']}%"])
        writer.writerow([])
        
        # Weekly breakdown
        writer.writerow(['WEEKLY BREAKDOWN'])
        writer.writerow(['Week', 'Events', 'Confirmed', 'Tentative', 'Revenue', 'Attendees'])
        
        for week_name, week_data in monthly_data['weekly_summaries'].items():
            writer.writerow([
                week_name,
                week_data['events'],
                week_data['confirmed'],
                week_data['tentative'],
                f"${week_data['revenue']:,.2f}",
                week_data['attendees']
            ])
        
        writer.writerow([])
        
        # Top rooms
        writer.writerow(['TOP PERFORMING ROOMS'])
        writer.writerow(['Room', 'Events', 'Confirmed', 'Revenue', 'Attendees'])
        
        for room_name, room_data in monthly_data['top_rooms'][:10]:
            writer.writerow([
                room_name,
                room_data['events'],
                room_data['confirmed'],
                f"${room_data['revenue']:,.2f}",
                room_data['attendees']
            ])
        
        writer.writerow([])
        
        # Top clients
        writer.writerow(['TOP CLIENTS'])
        writer.writerow(['Client', 'Events', 'Confirmed', 'Revenue', 'Attendees'])
        
        for client_name, client_data in monthly_data['top_clients'][:10]:
            writer.writerow([
                client_name,
                client_data['events'],
                client_data['confirmed'],
                f"${client_data['revenue']:,.2f}",
                client_data['attendees']
            ])
        
        output.seek(0)
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'monthly_summary_{start_date.strftime("%Y-%m")}.csv'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export monthly summary CSV: {e}")
        raise

# ===============================
# PDF EXPORT FUNCTIONS
# ===============================

def export_daily_summary_pdf(daily_data, report_date):
    """Export daily summary as PDF file"""
    try:
        if not REPORTLAB_AVAILABLE:
            return export_daily_summary_csv(daily_data, report_date)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Build content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        story.append(Paragraph(f"<b>Daily Summary Report</b>", styles['Title']))
        story.append(Paragraph(f"<b>{report_date.strftime('%A, %B %d, %Y')}</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Summary statistics
        summary = daily_data['summary']
        story.append(Paragraph("<b>Summary Statistics</b>", styles['Heading2']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Events', str(summary['total_events'])],
            ['Confirmed Events', str(summary['confirmed_events'])],
            ['Tentative Events', str(summary['tentative_events'])],
            ['Total Revenue', f"${summary['total_revenue']:,.2f}"],
            ['Total Attendees', f"{summary['total_attendees']:,}"],
            ['Rooms in Use', str(summary['rooms_in_use'])]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Events by room
        story.append(Paragraph("<b>Events by Room</b>", styles['Heading2']))
        
        for room_name, events in daily_data['events_by_room'].items():
            if events:
                story.append(Paragraph(f"<b>{room_name}</b>", styles['Heading3']))
                
                events_data = [['Event', 'Client', 'PAX', 'Time', 'Revenue', 'Status']]
                
                for event in events:
                    events_data.append([
                        event['title'],
                        event['client_name'],
                        str(event['attendees']),
                        event['time_display'],
                        f"${event['total_price']:,.2f}",
                        event['status_display']
                    ])
                
                events_table = Table(events_data)
                events_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(events_table)
                story.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'daily_summary_{report_date.strftime("%Y-%m-%d")}.pdf'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export daily summary PDF: {e}")
        return export_daily_summary_csv(daily_data, report_date)

def export_weekly_summary_pdf(weekly_data, start_date, end_date):
    """Export weekly summary as PDF file"""
    try:
        if not REPORTLAB_AVAILABLE:
            return export_weekly_summary_csv(weekly_data, start_date, end_date)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Build content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        story.append(Paragraph(f"<b>Weekly Summary Report</b>", styles['Title']))
        story.append(Paragraph(f"<b>{start_date} to {end_date}</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Note about table format
        story.append(Paragraph("<i>Note: This is a simplified view. For the full table format, please use Excel export.</i>", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Summary
        summary = weekly_data['summary']
        story.append(Paragraph("<b>Weekly Summary</b>", styles['Heading2']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Events', str(summary['total_events'])],
            ['Total Revenue', f"${summary['total_revenue']:,.2f}"],
            ['Total Attendees', f"{summary['total_attendees']:,}"],
            ['Rooms with Events', str(summary['rooms_with_events'])],
            ['Average Daily Events', str(summary['average_daily_events'])]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Events by day
        for day in weekly_data['week_days']:
            day_events = []
            for room_name, room_data in weekly_data['room_schedule'].items():
                events = room_data['days'].get(day['date'], [])
                for event in events:
                    day_events.append({
                        'room': room_name,
                        'event': event
                    })
            
            if day_events:
                story.append(Paragraph(f"<b>{day['day_name']} - {day['date_display']}</b>", styles['Heading3']))
                
                day_data = [['Room', 'Client', 'PAX', 'Time', 'Revenue']]
                
                for item in day_events:
                    event = item['event']
                    day_data.append([
                        item['room'],
                        event['client_name'],
                        str(event['attendees']),
                        event['start_time'],
                        f"${event['total_price']:,.2f}"
                    ])
                
                day_table = Table(day_data)
                day_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(day_table)
                story.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'weekly_summary_{start_date}_to_{end_date}.pdf'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export weekly summary PDF: {e}")
        return export_weekly_summary_csv(weekly_data, start_date, end_date)

def export_monthly_summary_pdf(monthly_data, start_date, end_date):
    """Export monthly summary as PDF file"""
    try:
        if not REPORTLAB_AVAILABLE:
            return export_monthly_summary_csv(monthly_data, start_date, end_date)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Build content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        story.append(Paragraph(f"<b>Monthly Summary Report</b>", styles['Title']))
        story.append(Paragraph(f"<b>{monthly_data['month_name']}</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Overall summary
        summary = monthly_data['summary']
        story.append(Paragraph("<b>Overall Summary</b>", styles['Heading2']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Events', str(summary['total_events'])],
            ['Confirmed Events', str(summary['confirmed_events'])],
            ['Tentative Events', str(summary['tentative_events'])],
            ['Total Revenue', f"${summary['total_revenue']:,.2f}"],
            ['Total Attendees', f"{summary['total_attendees']:,}"],
            ['Conversion Rate', f"{summary['conversion_rate']}%"],
            ['Average Event Value', f"${summary['average_event_value']:,.2f}"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Top rooms
        if monthly_data['top_rooms']:
            story.append(Paragraph("<b>Top Performing Rooms</b>", styles['Heading2']))
            
            rooms_data = [['Room', 'Events', 'Confirmed', 'Revenue']]
            for room_name, room_data in monthly_data['top_rooms'][:10]:
                rooms_data.append([
                    room_name,
                    str(room_data['events']),
                    str(room_data['confirmed']),
                    f"${room_data['revenue']:,.2f}"
                ])
            
            rooms_table = Table(rooms_data)
            rooms_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(rooms_table)
            story.append(Spacer(1, 20))
        
        # Top clients
        if monthly_data['top_clients']:
            story.append(Paragraph("<b>Top Clients</b>", styles['Heading2']))
            
            clients_data = [['Client', 'Events', 'Confirmed', 'Revenue']]
            for client_name, client_data in monthly_data['top_clients'][:10]:
                clients_data.append([
                    client_name,
                    str(client_data['events']),
                    str(client_data['confirmed']),
                    f"${client_data['revenue']:,.2f}"
                ])
            
            clients_table = Table(clients_data)
            clients_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(clients_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'monthly_summary_{start_date.strftime("%Y-%m")}.pdf'
        )
        
    except Exception as e:
        print(f"❌ ERROR: Failed to export monthly summary PDF: {e}")
        return export_monthly_summary_csv(monthly_data, start_date, end_date)

# ===============================
# ERROR HANDLERS
# ===============================

@reports_bp.errorhandler(500)
def reports_internal_error(error):
    """Handle internal server errors in reports"""
    print(f"❌ Reports error: {error}")
    
    try:
        log_user_activity(
            ActivityTypes.ERROR_OCCURRED,
            f"Reports internal error: {str(error)}",
            resource_type='reports',
            status='failed'
        )
    except:
        pass
    
    flash('⚠️ An error occurred while generating the report. Please try again.', 'danger')
    return redirect(url_for('reports.reports'))

@reports_bp.errorhandler(404)
def reports_not_found(error):
    """Handle 404 errors in reports"""
    flash('⚠️ The requested report was not found.', 'warning')
    return redirect(url_for('reports.reports'))